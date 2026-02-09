#!/usr/bin/env python3
r"""
rekordbox_analyse.py

Baseline Rekordbox XML analyser, inspired by Library Dojo style reporting.

What it does (all optional, default is ALL):
- Overview stats (counts, duration, size, file types, bitrates, sample rates, BPM, key, genres, artists, labels, years)
- Data quality checks (missing fields, zero BPM, low bitrate, etc)
- Duplicate detection (by location, by artist+title, by signature)
- Playlist analysis (tree paths, counts, orphans, broken refs)
- Missing file checks and optional smart relink suggestions
- Optional embedded artwork scan (no art, placeholder art), generates .m3u8 playlists
- Optional Rekordbox vs Mixed In Key CSV comparison (missing tracks, BPM/key diffs)

Designed to stream parse large XML files (iterparse) and avoid loading it all into memory.

Usage:
  python rekordbox_analyse.py "C:\path\Rekordbox Collection.xml" --outdir "reports"
  python rekordbox_analyse.py "...xml" --music-root "D:\Music" --mik-csv "MIK.csv"
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Optional, Tuple
from urllib.parse import unquote, urlparse

import xml.etree.ElementTree as ET

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False

try:
    from mutagen import File as MutagenFile
    MUTAGEN_OK = True
except Exception:
    MUTAGEN_OK = False


# ---------------------------
# Helpers
# ---------------------------

def norm_ws(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip()

def safe_int(v: Any, default: int = 0) -> int:
    try:
        if v is None:
            return default
        if isinstance(v, (int,)):
            return int(v)
        s = str(v).strip()
        if not s:
            return default
        return int(float(s))
    except Exception:
        return default

def safe_float(v: Any, default: float = 0.0) -> float:
    try:
        if v is None:
            return default
        if isinstance(v, (float, int)):
            return float(v)
        s = str(v).strip()
        if not s:
            return default
        return float(s)
    except Exception:
        return default

def bytes_to_gb(n: int) -> float:
    return n / (1024**3)

def seconds_to_hhmmss(total: int) -> str:
    h = total // 3600
    m = (total % 3600) // 60
    s = total % 60
    return f"{h:02d}:{m:02d}:{s:02d}"

def fileurl_to_path(loc: str) -> str:
    """
    Rekordbox 'Location' is usually a file:// URL. Convert to a usable local path.
    Supports Windows and POSIX-ish file URLs.
    """
    if not loc:
        return ""
    loc = loc.strip()
    # Rekordbox sometimes stores as already-ish path
    if not loc.lower().startswith("file:"):
        return loc

    p = urlparse(loc)
    path = unquote(p.path or "")
    netloc = p.netloc or ""
    # Windows file urls often look like: file://localhost/D:/Music/Track.mp3
    if netloc and netloc.lower() not in ("localhost",):
        # network share, keep UNC style
        path = f"//{netloc}{path}"
    # Strip leading slash for drive letters: /D:/Music -> D:/Music
    if re.match(r"^/[A-Za-z]:/", path):
        path = path[1:]
    # Normalize slashes
    return path.replace("/", os.sep)

def sha1_bytes(b: bytes) -> str:
    return hashlib.sha1(b).hexdigest()

def sha1_file(fp: Path, chunk: int = 1024 * 1024) -> str:
    h = hashlib.sha1()
    with fp.open("rb") as f:
        while True:
            data = f.read(chunk)
            if not data:
                break
            h.update(data)
    return h.hexdigest()

def load_settings(settings_path: Optional[Path]) -> Dict[str, Any]:
    if not settings_path:
        return {}
    if not settings_path.exists():
        return {}
    try:
        return json.loads(settings_path.read_text(encoding="utf-8", errors="ignore"))
    except Exception:
        return {}

def ensure_outdir(outdir: Path) -> Path:
    outdir.mkdir(parents=True, exist_ok=True)
    return outdir

def write_csv(path: Path, rows: List[Dict[str, Any]], fieldnames: List[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fieldnames})


# ---------------------------
# Data models
# ---------------------------

@dataclass
class Track:
    track_id: str
    name: str
    artist: str
    album: str
    remixer: str
    mix: str
    label: str
    genre: str
    year: int
    date_added: str
    bpm: float
    key: str
    kind: str
    size: int
    duration: int
    bitrate: int
    samplerate: int
    rating: str
    playcount: int
    comments: str
    location_url: str
    location_path: str

    def title_key(self) -> str:
        return f"{norm_key(self.artist)}|{norm_key(self.name)}"

def norm_key(s: str) -> str:
    s = (s or "").lower()
    s = re.sub(r"\(.*?\)", "", s)  # remove bracketed bits
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return norm_ws(s)

@dataclass
class Playlist:
    path: str
    name: str
    node_type: str
    entries: int
    track_keys: List[str]


# ---------------------------
# XML parsing
# ---------------------------

def iter_tracks(xml_path: Path) -> Iterator[Track]:
    # Track elements live under COLLECTION/TRACK
    ctx = ET.iterparse(str(xml_path), events=("start", "end"))
    _, root = next(ctx)  # type: ignore[misc]
    in_collection = False

    for event, elem in ctx:
        tag = elem.tag.upper()
        if event == "start" and tag == "COLLECTION":
            in_collection = True
        elif event == "end" and tag == "COLLECTION":
            in_collection = False

        if event == "end" and in_collection and tag == "TRACK":
            a = elem.attrib
            loc = a.get("Location", "") or a.get("location", "")
            loc_path = fileurl_to_path(loc)
            t = Track(
                track_id=str(a.get("TrackID", "")),
                name=str(a.get("Name", "")),
                artist=str(a.get("Artist", "")),
                album=str(a.get("Album", "")),
                remixer=str(a.get("Remixer", "")),
                mix=str(a.get("Mix", "")),
                label=str(a.get("Label", "")),
                genre=str(a.get("Genre", "")),
                year=safe_int(a.get("Year", 0), 0),
                date_added=str(a.get("DateAdded", "")),
                bpm=safe_float(a.get("AverageBpm", 0.0), 0.0),
                key=str(a.get("Tonality", "")),
                kind=str(a.get("Kind", "")),
                size=safe_int(a.get("Size", 0), 0),
                duration=safe_int(a.get("TotalTime", 0), 0),
                bitrate=safe_int(a.get("BitRate", 0), 0),
                samplerate=safe_int(a.get("SampleRate", 0), 0),
                rating=str(a.get("Rating", "")),
                playcount=safe_int(a.get("PlayCount", 0), 0),
                comments=str(a.get("Comments", "")),
                location_url=loc,
                location_path=loc_path,
            )
            yield t
            elem.clear()
            root.clear()

def parse_playlists(xml_path: Path) -> List[Playlist]:
    playlists: List[Playlist] = []
    ctx = ET.iterparse(str(xml_path), events=("start", "end"))
    _, root = next(ctx)  # type: ignore[misc]
    in_playlists = False
    stack: List[Tuple[str, Dict[str, str]]] = []
    current_tracks: List[str] = []

    def stack_path() -> str:
        names = [n for n, _a in stack if n and n != "ROOT"]
        return "/".join(names)

    for event, elem in ctx:
        tag = elem.tag.upper()
        if event == "start" and tag == "PLAYLISTS":
            in_playlists = True
        elif event == "end" and tag == "PLAYLISTS":
            in_playlists = False

        if not in_playlists:
            continue

        if tag == "NODE" and event == "start":
            name = elem.attrib.get("Name", "")
            stack.append((name, dict(elem.attrib)))
            current_tracks = []

        elif tag == "TRACK" and event == "start":
            k = elem.attrib.get("Key")
            if k:
                current_tracks.append(str(k))

        elif tag == "NODE" and event == "end":
            if stack:
                name, attrs = stack.pop()
                node_type = str(attrs.get("Type", ""))
                entries = safe_int(attrs.get("Entries", 0), 0)
                # If it contains track keys, treat it as a playlist-like node
                if current_tracks or entries:
                    full_path = stack_path()
                    full = f"{full_path}/{name}".strip("/") if name else full_path
                    playlists.append(Playlist(
                        path=full,
                        name=name,
                        node_type=node_type,
                        entries=entries if entries else len(current_tracks),
                        track_keys=list(current_tracks),
                    ))
            current_tracks = []
            elem.clear()
            root.clear()

    # De-dupe by path
    seen = set()
    uniq = []
    for p in playlists:
        if p.path in seen:
            continue
        seen.add(p.path)
        uniq.append(p)
    return uniq


# ---------------------------
# MIK CSV parsing (best-effort)
# ---------------------------

def read_mik_csv(mik_csv: Path) -> Dict[str, Dict[str, Any]]:
    """
    Returns dict keyed by normalized file path.
    We try to detect common columns: path, bpm, key.
    """
    with mik_csv.open("r", encoding="utf-8", errors="ignore", newline="") as f:
        reader = csv.DictReader(f)
        cols = [c or "" for c in (reader.fieldnames or [])]
        lower = [c.lower() for c in cols]

        def find_col(cands: List[str]) -> Optional[str]:
            for cand in cands:
                for c in cols:
                    if c.lower() == cand.lower():
                        return c
            # contains match
            for cand in cands:
                for c in cols:
                    if cand.lower() in c.lower():
                        return c
            return None

        path_col = find_col(["path", "file", "filename", "location", "filepath", "file path"])
        bpm_col = find_col(["bpm", "tempo"])
        key_col = find_col(["key", "camelot", "tonality"])

        data: Dict[str, Dict[str, Any]] = {}
        for row in reader:
            rawp = (row.get(path_col, "") if path_col else "") or ""
            p = rawp.replace("/", os.sep).replace("\\\\", "\\").strip()
            if not p:
                continue
            k = p.lower()
            data[k] = {
                "path": p,
                "bpm": safe_float(row.get(bpm_col, ""), 0.0) if bpm_col else 0.0,
                "key": (row.get(key_col, "") if key_col else "") or "",
                "row": row,
            }
        return data


# ---------------------------
# Artwork scan
# ---------------------------

def extract_embedded_art_sha1(fp: Path) -> Tuple[bool, Optional[str], Optional[int]]:
    """
    Returns (has_art, sha1, bytes_len)
    """
    if not MUTAGEN_OK:
        return (False, None, None)
    try:
        mf = MutagenFile(str(fp))
        if mf is None:
            return (False, None, None)

        # FLAC
        if hasattr(mf, "pictures") and getattr(mf, "pictures"):
            pics = getattr(mf, "pictures")
            if pics:
                b = pics[0].data
                return (True, sha1_bytes(b), len(b))

        # MP3 ID3 APIC
        if hasattr(mf, "tags") and mf.tags:
            tags = mf.tags
            # mutagen.id3
            for k in tags.keys():
                if str(k).startswith("APIC"):
                    apic = tags.get(k)
                    b = getattr(apic, "data", None)
                    if b:
                        return (True, sha1_bytes(b), len(b))

        # AIFF/WAV may store in tags too (less common)
        return (False, None, None)
    except Exception:
        return (False, None, None)


# ---------------------------
# Excel writing
# ---------------------------

def autosize(ws: Worksheet, max_width: int = 60) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_width, max(10, max_len + 2))

def write_excel_report(
    xlsx_path: Path,
    overview_rows: List[Tuple[str, Any]],
    issues: List[Dict[str, Any]],
    dup_groups: List[Dict[str, Any]],
    playlist_rows: List[Dict[str, Any]],
    orphan_rows: List[Dict[str, Any]],
    broken_rows: List[Dict[str, Any]],
    missing_file_rows: List[Dict[str, Any]],
    artwork_rows: List[Dict[str, Any]],
    mik_rows: List[Dict[str, Any]],
) -> None:
    if not OPENPYXL_OK:
        return
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def add_sheet(name: str, headers: List[str], rows: List[Dict[str, Any]]):
        ws = wb.create_sheet(title=name[:31])
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        autosize(ws)

    # Overview
    ws = wb.create_sheet(title="Overview")
    ws.append(["Metric", "Value"])
    for k, v in overview_rows:
        ws.append([k, v])
    autosize(ws)

    add_sheet("Quality Issues",
              ["Issue", "TrackID", "Artist", "Title", "Genre", "Year", "BPM", "Key", "BitRate", "SampleRate", "Kind", "Location"],
              issues)

    add_sheet("Duplicates",
              ["DuplicateType", "GroupKey", "Count", "TrackIDs", "Artists", "Titles", "Locations"],
              dup_groups)

    add_sheet("Playlists",
              ["PlaylistPath", "NodeType", "Entries", "ActualTrackKeys"],
              playlist_rows)

    add_sheet("Orphans",
              ["TrackKey", "TrackID", "Artist", "Title", "Location"],
              orphan_rows)

    add_sheet("Broken Playlist Refs",
              ["PlaylistPath", "MissingTrackKey"],
              broken_rows)

    add_sheet("Missing Files",
              ["TrackID", "Artist", "Title", "Location", "Exists", "RelinkSuggestion"],
              missing_file_rows)

    add_sheet("Artwork Scan",
              ["TrackID", "Artist", "Title", "Location", "HasArt", "ArtBytes", "ArtSha1", "IsPlaceholder"],
              artwork_rows)

    add_sheet("MIK Compare",
              ["Status", "TrackID", "Artist", "Title", "RB_Path", "MIK_Path", "RB_BPM", "MIK_BPM", "RB_Key", "MIK_Key"],
              mik_rows)

    wb.save(str(xlsx_path))


# ---------------------------
# Main analysis
# ---------------------------

def analyse(
    xml_path: Path,
    outdir: Path,
    music_root: Optional[Path],
    mik_csv: Optional[Path],
    settings_path: Optional[Path],
    low_bitrate_mp3: int = 320,
    do_artwork: bool = True,
    do_missing_files: bool = True,
) -> None:
    outdir = ensure_outdir(outdir)

    settings = load_settings(settings_path)
    placeholder_fp = None
    placeholder_sha1 = None
    try:
        ph = settings.get("discogs", {}).get("placeholder_image", "")
        if ph:
            placeholder_fp = Path(ph)
            if placeholder_fp.exists():
                placeholder_sha1 = sha1_file(placeholder_fp)
    except Exception:
        placeholder_fp = None
        placeholder_sha1 = None

    tracks: List[Track] = []
    tracks_by_key: Dict[str, Track] = {}

    print(f"[INFO] Parsing tracks (stream): {xml_path}")
    for t in iter_tracks(xml_path):
        tracks.append(t)
        # Rekordbox playlist uses TRACK Key which is TrackID
        if t.track_id:
            tracks_by_key[str(t.track_id)] = t

    print(f"[OK] Tracks loaded: {len(tracks)}")

    # Overview stats
    total_size = sum(t.size for t in tracks)
    total_dur = sum(t.duration for t in tracks)
    kinds = Counter((t.kind or "").strip() for t in tracks)
    genres = Counter(norm_ws(t.genre) for t in tracks if (t.genre or "").strip())
    artists = Counter(norm_ws(t.artist) for t in tracks if (t.artist or "").strip())
    labels = Counter(norm_ws(t.label) for t in tracks if (t.label or "").strip())
    years = Counter(t.year for t in tracks if t.year)

    bitrates = [t.bitrate for t in tracks if t.bitrate]
    samplerates = [t.samplerate for t in tracks if t.samplerate]
    bpms = [t.bpm for t in tracks if t.bpm]
    keys = Counter((t.key or "").strip() for t in tracks if (t.key or "").strip())

    def stat_line(label: str, value: Any) -> Tuple[str, Any]:
        return (label, value)

    overview: List[Tuple[str, Any]] = [
        stat_line("Total tracks", len(tracks)),
        stat_line("Total duration (hh:mm:ss)", seconds_to_hhmmss(total_dur)),
        stat_line("Total size (GB)", round(bytes_to_gb(total_size), 2)),
        stat_line("Unique artists", len(artists)),
        stat_line("Unique labels", len(labels)),
        stat_line("Unique genres", len(genres)),
    ]
    if bpms:
        overview += [
            stat_line("BPM min", round(min(bpms), 2)),
            stat_line("BPM max", round(max(bpms), 2)),
            stat_line("BPM avg", round(sum(bpms) / len(bpms), 2)),
        ]
    if bitrates:
        overview += [
            stat_line("Bitrate min", min(bitrates)),
            stat_line("Bitrate max", max(bitrates)),
            stat_line("Bitrate avg", round(sum(bitrates) / len(bitrates), 2)),
        ]
    if samplerates:
        overview += [
            stat_line("SampleRate min", min(samplerates)),
            stat_line("SampleRate max", max(samplerates)),
        ]

    # Add top lists
    def top(counter: Counter, n: int = 15) -> str:
        items = counter.most_common(n)
        return ", ".join([f"{k} ({v})" for k, v in items if str(k).strip()])

    overview += [
        stat_line("Top genres", top(genres)),
        stat_line("Top artists", top(artists)),
        stat_line("Top labels", top(labels)),
        stat_line("Top keys", top(keys)),
        stat_line("Top years", ", ".join([f"{y} ({c})" for y, c in years.most_common(15)])),
        stat_line("File types", top(kinds)),
    ]

    # Quality issues
    issues: List[Dict[str, Any]] = []
    for t in tracks:
        def add(issue: str):
            issues.append({
                "Issue": issue,
                "TrackID": t.track_id,
                "Artist": t.artist,
                "Title": t.name,
                "Genre": t.genre,
                "Year": t.year,
                "BPM": t.bpm,
                "Key": t.key,
                "BitRate": t.bitrate,
                "SampleRate": t.samplerate,
                "Kind": t.kind,
                "Location": t.location_path,
            })

        if not t.artist.strip():
            add("Missing Artist")
        if not t.name.strip():
            add("Missing Title")
        if not t.genre.strip():
            add("Missing Genre")
        if not t.label.strip():
            add("Missing Label")
        if t.year == 0:
            add("Missing Year")
        if t.bpm <= 0:
            add("Missing/Zero BPM")
        if not t.key.strip():
            add("Missing Key")
        if t.kind.lower().startswith("mp3") and t.bitrate and t.bitrate < low_bitrate_mp3:
            add(f"Low MP3 bitrate (<{low_bitrate_mp3})")
        if "  " in (t.name or "") or "  " in (t.artist or ""):
            add("Double spaces in Artist/Title")
        if (t.name or "").endswith(" ") or (t.artist or "").endswith(" "):
            add("Trailing space in Artist/Title")

    # Duplicates
    dup_groups: List[Dict[str, Any]] = []

    def add_dup_group(dup_type: str, group_key: str, group: List[Track]):
        dup_groups.append({
            "DuplicateType": dup_type,
            "GroupKey": group_key,
            "Count": len(group),
            "TrackIDs": ", ".join([g.track_id for g in group]),
            "Artists": ", ".join([g.artist for g in group]),
            "Titles": ", ".join([g.name for g in group]),
            "Locations": ", ".join([g.location_path for g in group]),
        })

    # by location
    by_loc = defaultdict(list)
    for t in tracks:
        k = (t.location_path or "").lower()
        if k:
            by_loc[k].append(t)
    for k, grp in by_loc.items():
        if len(grp) > 1:
            add_dup_group("Same Location", k, grp)

    # by artist+title
    by_title = defaultdict(list)
    for t in tracks:
        k = t.title_key()
        if k.strip("|"):
            by_title[k].append(t)
    for k, grp in by_title.items():
        if len(grp) > 1:
            add_dup_group("Same Artist+Title", k, grp)

    # signature (artist+title+duration+rounded bpm)
    by_sig = defaultdict(list)
    for t in tracks:
        rbpm = int(round(t.bpm)) if t.bpm else 0
        k = f"{t.title_key()}|{t.duration}|{rbpm}"
        by_sig[k].append(t)
    for k, grp in by_sig.items():
        if len(grp) > 1:
            add_dup_group("Signature Match", k, grp)

    # Playlist analysis
    print("[INFO] Parsing playlists...")
    playlists = parse_playlists(xml_path)
    playlist_rows = [{
        "PlaylistPath": p.path,
        "NodeType": p.node_type,
        "Entries": p.entries,
        "ActualTrackKeys": len(p.track_keys),
    } for p in playlists]

    used_keys = set()
    broken_rows: List[Dict[str, Any]] = []
    for p in playlists:
        for k in p.track_keys:
            used_keys.add(k)
            if k not in tracks_by_key:
                broken_rows.append({"PlaylistPath": p.path, "MissingTrackKey": k})

    orphan_rows: List[Dict[str, Any]] = []
    for k, t in tracks_by_key.items():
        if k not in used_keys:
            orphan_rows.append({
                "TrackKey": k,
                "TrackID": t.track_id,
                "Artist": t.artist,
                "Title": t.name,
                "Location": t.location_path,
            })

    # Missing file checks (+ smart relink)
    missing_file_rows: List[Dict[str, Any]] = []
    relink_index: Dict[str, str] = {}
    if do_missing_files:
        if music_root and music_root.exists():
            print(f"[INFO] Building relink index under: {music_root}")
            # filename -> full path (first match)
            for fp in music_root.rglob("*"):
                if fp.is_file():
                    relink_index.setdefault(fp.name.lower(), str(fp))
        else:
            music_root = None

        print("[INFO] Checking file existence...")
        for t in tracks:
            p = Path(t.location_path) if t.location_path else None
            exists = bool(p and p.exists())
            relink = ""
            if not exists and relink_index and p:
                relink = relink_index.get(p.name.lower(), "")
            missing_file_rows.append({
                "TrackID": t.track_id,
                "Artist": t.artist,
                "Title": t.name,
                "Location": t.location_path,
                "Exists": "Y" if exists else "N",
                "RelinkSuggestion": relink,
            })

    # Artwork scan
    artwork_rows: List[Dict[str, Any]] = []
    no_art_paths: List[str] = []
    placeholder_paths: List[str] = []
    if do_artwork:
        if not MUTAGEN_OK:
            print("[WARN] mutagen not installed, skipping embedded artwork scan.")
        else:
            print("[INFO] Scanning embedded artwork (can take time on large libraries)...")
            for t in tracks:
                p = Path(t.location_path) if t.location_path else None
                if not p or not p.exists():
                    continue
                has_art, art_sha1, art_len = extract_embedded_art_sha1(p)
                is_placeholder = (placeholder_sha1 is not None and art_sha1 == placeholder_sha1) if art_sha1 else False
                artwork_rows.append({
                    "TrackID": t.track_id,
                    "Artist": t.artist,
                    "Title": t.name,
                    "Location": t.location_path,
                    "HasArt": "Y" if has_art else "N",
                    "ArtBytes": art_len or "",
                    "ArtSha1": art_sha1 or "",
                    "IsPlaceholder": "Y" if is_placeholder else "N",
                })
                if not has_art:
                    no_art_paths.append(t.location_path)
                elif is_placeholder:
                    placeholder_paths.append(t.location_path)

    # Write issue playlists (.m3u8)
    def write_m3u8(name: str, paths: List[str]):
        if not paths:
            return
        fp = outdir / f"{name}.m3u8"
        with fp.open("w", encoding="utf-8") as f:
            f.write("#EXTM3U\n")
            for p in paths:
                if p:
                    f.write(p + "\n")
        print(f"[OK] Wrote playlist: {fp}")

    write_m3u8("no_artwork", no_art_paths)
    write_m3u8("placeholder_artwork", placeholder_paths)

    # MIK compare
    mik_rows: List[Dict[str, Any]] = []
    if mik_csv and mik_csv.exists():
        print(f"[INFO] Loading MIK CSV: {mik_csv}")
        mik = read_mik_csv(mik_csv)
        rb_paths = { (t.location_path or "").lower(): t for t in tracks if t.location_path }
        # missing in MIK
        for pth, t in rb_paths.items():
            if pth not in mik:
                mik_rows.append({
                    "Status": "Missing in MIK",
                    "TrackID": t.track_id, "Artist": t.artist, "Title": t.name,
                    "RB_Path": t.location_path, "MIK_Path": "",
                    "RB_BPM": t.bpm, "MIK_BPM": "",
                    "RB_Key": t.key, "MIK_Key": "",
                })
        # missing in RB
        for pth, m in mik.items():
            if pth not in rb_paths:
                mik_rows.append({
                    "Status": "Missing in Rekordbox",
                    "TrackID": "", "Artist": "", "Title": "",
                    "RB_Path": "", "MIK_Path": m.get("path",""),
                    "RB_BPM": "", "MIK_BPM": m.get("bpm",""),
                    "RB_Key": "", "MIK_Key": m.get("key",""),
                })
        # diffs
        for pth, t in rb_paths.items():
            m = mik.get(pth)
            if not m:
                continue
            mbpm = safe_float(m.get("bpm", 0.0), 0.0)
            mkey = (m.get("key", "") or "").strip()
            status = []
            if mbpm and t.bpm and abs(mbpm - t.bpm) >= 0.75:
                status.append("BPM diff")
            if mkey and t.key and norm_key(mkey) != norm_key(t.key):
                status.append("Key diff")
            if status:
                mik_rows.append({
                    "Status": ", ".join(status),
                    "TrackID": t.track_id, "Artist": t.artist, "Title": t.name,
                    "RB_Path": t.location_path, "MIK_Path": m.get("path",""),
                    "RB_BPM": t.bpm, "MIK_BPM": mbpm,
                    "RB_Key": t.key, "MIK_Key": mkey,
                })

    # Write outputs
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    summary_txt = outdir / f"rekordbox_report_{ts}.txt"
    xlsx_path = outdir / f"rekordbox_report_{ts}.xlsx"

    print(f"[INFO] Writing summary: {summary_txt}")
    with summary_txt.open("w", encoding="utf-8") as f:
        f.write("Baseline Rekordbox XML Report\n")
        f.write(f"XML: {xml_path}\n")
        f.write(f"Generated: {datetime.now().isoformat(timespec='seconds')}\n\n")
        for k, v in overview:
            f.write(f"{k}: {v}\n")
        f.write("\nCounts\n")
        f.write(f"Quality issues: {len(issues)}\n")
        f.write(f"Duplicate groups: {len(dup_groups)}\n")
        f.write(f"Playlists: {len(playlists)}\n")
        f.write(f"Orphans: {len(orphan_rows)}\n")
        f.write(f"Broken playlist refs: {len(broken_rows)}\n")
        if do_missing_files:
            missing = sum(1 for r in missing_file_rows if r.get("Exists") == "N")
            f.write(f"Missing files: {missing}\n")
        if do_artwork and MUTAGEN_OK:
            f.write(f"Artwork scanned rows: {len(artwork_rows)}\n")
            f.write(f"No artwork: {len(no_art_paths)}\n")
            f.write(f"Placeholder artwork: {len(placeholder_paths)}\n")
        if mik_rows:
            f.write(f"MIK compare rows: {len(mik_rows)}\n")

    # Excel
    if OPENPYXL_OK:
        print(f"[INFO] Writing Excel: {xlsx_path}")
        write_excel_report(
            xlsx_path,
            overview,
            issues,
            dup_groups,
            playlist_rows,
            orphan_rows,
            broken_rows,
            missing_file_rows,
            artwork_rows,
            mik_rows,
        )
    else:
        print("[WARN] openpyxl not installed, skipping Excel output.")

    # Also dump CSVs for easy grepping
    print("[INFO] Writing CSV extracts...")
    write_csv(outdir / f"quality_issues_{ts}.csv",
              issues,
              ["Issue", "TrackID", "Artist", "Title", "Genre", "Year", "BPM", "Key", "BitRate", "SampleRate", "Kind", "Location"])
    write_csv(outdir / f"duplicates_{ts}.csv",
              dup_groups,
              ["DuplicateType", "GroupKey", "Count", "TrackIDs", "Artists", "Titles", "Locations"])
    write_csv(outdir / f"playlists_{ts}.csv",
              playlist_rows,
              ["PlaylistPath", "NodeType", "Entries", "ActualTrackKeys"])
    write_csv(outdir / f"orphans_{ts}.csv",
              orphan_rows,
              ["TrackKey", "TrackID", "Artist", "Title", "Location"])
    write_csv(outdir / f"broken_playlist_refs_{ts}.csv",
              broken_rows,
              ["PlaylistPath", "MissingTrackKey"])
    if missing_file_rows:
        write_csv(outdir / f"missing_files_{ts}.csv",
                  missing_file_rows,
                  ["TrackID", "Artist", "Title", "Location", "Exists", "RelinkSuggestion"])
    if artwork_rows:
        write_csv(outdir / f"artwork_scan_{ts}.csv",
                  artwork_rows,
                  ["TrackID", "Artist", "Title", "Location", "HasArt", "ArtBytes", "ArtSha1", "IsPlaceholder"])
    if mik_rows:
        write_csv(outdir / f"mik_compare_{ts}.csv",
                  mik_rows,
                  ["Status", "TrackID", "Artist", "Title", "RB_Path", "MIK_Path", "RB_BPM", "MIK_BPM", "RB_Key", "MIK_Key"])

    print("\n[OK] Done.")
    print(f"Summary: {summary_txt}")
    if OPENPYXL_OK:
        print(f"Excel:   {xlsx_path}")
    print(f"Outdir:  {outdir}")


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Baseline Rekordbox XML analyser")
    p.add_argument("xml", help="Path to Rekordbox exported XML (Collection.xml)")
    p.add_argument("--outdir", default="", help="Output folder for reports (default: ./baseline_work/rekordbox_reports)")
    p.add_argument("--music-root", default="", help="Optional folder root to index for relink suggestions")
    p.add_argument("--mik-csv", default="", help="Optional Mixed In Key CSV export to compare")
    p.add_argument("--settings", default="", help="Optional baseline settings json (to find placeholder image)")
    p.add_argument("--low-bitrate-mp3", type=int, default=320, help="Threshold for MP3 bitrate flag")
    p.add_argument("--no-artwork-scan", action="store_true", help="Disable embedded artwork scan")
    p.add_argument("--no-missing-files", action="store_true", help="Disable file existence checks")
    return p


def main() -> None:
    ap = build_arg_parser()
    args = ap.parse_args()

    xml_path = Path(args.xml).expanduser()
    if not xml_path.exists():
        print(f"[ERROR] XML not found: {xml_path}")
        sys.exit(2)

    default_out = Path.cwd() / "baseline_work" / "rekordbox_reports"
    outdir = Path(args.outdir).expanduser() if args.outdir else default_out

    music_root = Path(args.music_root).expanduser() if args.music_root else None
    mik_csv = Path(args.mik_csv).expanduser() if args.mik_csv else None
    settings_path = Path(args.settings).expanduser() if args.settings else None

    analyse(
        xml_path=xml_path,
        outdir=outdir,
        music_root=music_root,
        mik_csv=mik_csv,
        settings_path=settings_path,
        low_bitrate_mp3=args.low_bitrate_mp3,
        do_artwork=not args.no_artwork_scan,
        do_missing_files=not args.no_missing_files,
    )


if __name__ == "__main__":
    main()
